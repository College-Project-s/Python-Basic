import openpyxl
from openpyxl import load_workbook
import os

# Nama file excel
file_excel = 'data_kendaraan_format.xlsx'

# Cek apakah file ada
if not os.path.exists(file_excel):
    print("❌ File Excel tidak ditemukan. Jalankan skrip pembuat file terlebih dahulu.")
    exit()

# Fungsi tambah data ke sheet
def tambah_data(sheet_name, data):
    wb = load_workbook(file_excel)
    ws = wb[sheet_name]
    ws.append(data)
    wb.save(file_excel)

# Perulangan input data
while True:
    print("\nMasukkan data kendaraan:")
    merk = input("Merk: ")
    tahun = input("Tahun: ")
    warna = input("Warna: ")
    data = [merk, tahun, warna]

    print("\nPilih format penyimpanan:")
    print("1. List")
    print("2. Tuple")
    print("3. Set (data unik)")

    pilihan = input("Pilihan (1/2/3): ")

    if pilihan == "1":
        tambah_data("Kendaraan_List", data)
    elif pilihan == "2":
        tambah_data("Kendaraan_Tuple", tuple(data))
    elif pilihan == "3":
        wb = load_workbook(file_excel)
        ws = wb["Kendaraan_Set"]
        data_set = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_set.add(tuple(row))
        data_set.add(tuple(data))
        # Clear dulu isi lama
        ws.delete_rows(2, ws.max_row)
        for item in data_set:
            ws.append(list(item))
        wb.save(file_excel)
    else:
        print("❌ Pilihan tidak valid.")

    lanjut = input("Ingin input data lagi? (ya/tidak): ").lower()
    if lanjut != "ya":
        break

# Tampilkan semua data
wb = load_workbook(file_excel)

print("\n📄 Isi semua sheet:")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\n📘 Sheet: {sheet}")
    for row in ws.iter_rows(values_only=True):
        print(row)
